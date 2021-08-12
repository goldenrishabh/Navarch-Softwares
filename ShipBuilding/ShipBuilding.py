#All rights reserved by NA20 OP Batch
#Illegal distribution may lead to rick roll

from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook("ShipDrawing.xlsx")
print("Sheet Loaded!!\n")
sheet = workbook.active

scale = float(input("Enter scale value : "))
width = float(input("Enter paper width (cm) : "))
length = float(input("Enter paper length (longest dimension)(cm) : "))
margin = float(input("Enter paper margin(one side )(cm) :"))

def myround(y, base=5):
    if ((int(y)) - (int(y))% 5) > y:
        return (int(y)) - (int(y))% 5
    return (int(y)+5) - (int(y)+5)% 5

sheet["A28"].value = "Paper Size"
sheet["A29"].value = "width"
sheet["A30"].value = "height"
sheet["B29"].value = width
sheet["B30"].value = length
sheet["D28"].value = "Scaling factor"
sheet["E28"].value = scale
sheet["D29"].value = "Station length"
sheet["E29"].value = 110.70 * scale /20
sheet["D30"].value = "Aft Length"
sheet["E30"].value = (sheet["E29"].value * 8) + (3.6*scale)
sheet["D31"].value = "Forward Length"
sheet["E31"].value =(sheet["E29"].value * 5+(0.9 *scale))
sheet["G28"].value = "Drawing scale"
sheet["H28"].value = sheet["E30"].value /((length-2*margin)/100) 
sheet["G29"].value = "Final ratio"
a = myround(sheet["H28"].value)
sheet["H29"].value = "1 : "+ str(a)
sheet["E29"].value = 110.70 * scale /20 *100/a
sheet["D30"].value = "Aft Length"
sheet["E30"].value = sheet["E30"].value*100/ a 
sheet["D31"].value = "Forward Length"
sheet["E31"].value = sheet["E31"].value* 100/a
sheet["J28"].value = "LOA"
sheet["J29"].value = "LBP"
sheet["J30"].value = "B"
sheet["J31"].value = "D"
sheet["J32"].value = "T"

sheet["K28"].value = 117.04 * scale *100 / a
sheet["K29"].value = 110.70 * scale *100 / a
sheet["K30"].value = 19.86 * scale  * 100 /a
sheet["K31"].value = 10.18 * scale *100 / a
sheet["K32"].value = 5.10 * scale  * 100/a

sheet["M28"].value = "Body Plan"
sheet["M29"].value = "Butock section"
sheet["M30"].value = "Total Buttock Section"
sheet["M31"].value = "Waterline Spacing"
sheet["M32"].value = "Estimated Draft"
sheet["M33"].value = "SuperStructure"

sheet["N29"].value = 19.86 * scale /8 *100/a
sheet["N30"].value = 19.86 * scale *100/a
sheet["N31"].value = 1 * scale *100/a
sheet["N32"].value = 1 * scale * 10 * 100 /a
sheet["N33"].value = 5 * sheet["N31"].value

for row in range(2,27):
    for column in range(2,17):
        if type(sheet.cell(row,column).value)==float:
             sheet.cell(row,column).value /= a/100
             sheet.cell(row,column).value *= scale
        elif type(sheet.cell(row,column).value)==int :
            sheet.cell(row,column).value = float(sheet.cell(row,column).value)
            sheet.cell(row,column).value /= a/100 
            sheet.cell(row,column).value *= scale
workbook.save("ShipBuildingCustom.xlsx")
print("Excel saved as ShipBuildingCustom")

